# ============================================================
# 3_excel.py — Excel avec pandas et openpyxl
# ============================================================
# Deux outils complementaires :
#   - pandas : lire/ecrire la data
#   - openpyxl : mettre en forme (couleurs, fonts, fusions...)
# ============================================================

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# 1. Lire un Excel avec pandas (le plus simple)
# ============================================================
df = pd.read_excel("data/commandes.xlsx")
print("Apercu Excel :")
print(df.head())
print()


# ============================================================
# 2. Faire un calcul rapide puis exporter
# ============================================================
total_par_produit = df.groupby("produit")["montant"].sum()
total_par_produit.to_excel("data/totaux.xlsx")


# ============================================================
# 3. Plusieurs onglets dans un meme fichier
# ============================================================
with pd.ExcelWriter("data/rapport_multi.xlsx") as writer:
    df.to_excel(writer, sheet_name="Donnees brutes", index=False)
    total_par_produit.to_excel(writer, sheet_name="Synthese")

    # On peut ajouter d'autres tableaux sur d'autres onglets
    top10 = df.nlargest(10, "montant")
    top10.to_excel(writer, sheet_name="Top 10", index=False)


# ============================================================
# 4. Mettre en forme avec openpyxl
# ============================================================
# On reprend le fichier qu'on vient de creer et on l'embellit.
wb = load_workbook("data/rapport_multi.xlsx")
ws = wb["Donnees brutes"]

# Styliser l'en-tete
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="2F75B5")

for cell in ws[1]:                   # ws[1] = premiere ligne
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Elargir les colonnes
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 2

# Geler la premiere ligne (figer en haut)
ws.freeze_panes = "A2"

wb.save("data/rapport_multi.xlsx")
print("data/rapport_multi.xlsx genere avec mise en forme")


# ============================================================
# Astuce : feuille protegee, formules, conditional formatting...
# openpyxl peut TOUT faire. Doc officielle :
# https://openpyxl.readthedocs.io/
# ============================================================
